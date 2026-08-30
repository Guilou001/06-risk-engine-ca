"""Le classeur, vérifié sur ce qu'il contient vraiment et non sur le fait qu'il s'écrive."""

import re
import zipfile

import pandas as pd
import pytest

from rke.classeur import construire


@pytest.fixture
def classeur(tmp_path):
    backtests = pd.DataFrame(
        {"n_jours": [5476, 5476], "depassements": [89, 69]},
        index=["historique", "fhs"])
    feux = {"fhs": pd.DataFrame({"n_jours": [250, 252], "depassements": [4, 11],
                                 "zone": ["vert", "rouge"]}, index=[2019, 2020])}
    return construire(backtests, feux, tmp_path / "essai.xlsx")


def _formules(chemin) -> list[str]:
    archive = zipfile.ZipFile(chemin)
    sortie = []
    for nom in archive.namelist():
        if "worksheets/sheet" in nom:
            sortie += re.findall(r"<f>([^<]+)</f>", archive.read(nom).decode("utf-8", "replace"))
    return sortie


def test_le_classeur_porte_ses_trois_feuilles(classeur):
    from openpyxl import load_workbook

    noms = load_workbook(classeur).sheetnames
    assert noms == ["Notice", "Kupiec", "Feux de Bâle"]


def test_les_calculs_sont_des_formules_et_non_des_resultats_colles(classeur):
    """C'est toute la raison d'être du classeur : un lecteur doit pouvoir changer une case et voir
    le résultat bouger. Un chiffre collé ne bouge pas."""
    formules = _formules(classeur)
    # six formules par modèle testé, trois par année de feu tricolore : deux modèles et deux années
    # font douze plus six, soit dix-huit
    assert len(formules) == 6 * 2 + 3 * 2
    assert any(f.startswith("CHIDIST(") for f in formules)
    assert any("LN(" in f for f in formules)


def test_la_fonction_du_khi_deux_est_ecrite_sous_sa_forme_ancienne(classeur):
    """Le piège du format : les fonctions ajoutées après Excel 2007 doivent porter un préfixe
    technique dans le fichier, sans quoi Excel affiche une erreur de nom. La forme ancienne, elle,
    fonctionne partout."""
    formules = _formules(classeur)
    assert not any("CHISQ.DIST" in f for f in formules)
    assert any("CHIDIST" in f for f in formules)


def test_les_seuils_de_bale_sont_mis_a_l_echelle_du_nombre_de_jours(classeur):
    """Les seuils de 5 et 10 valent pour 250 jours. Une année de 252 jours doit voir ses seuils
    remonter, et la formule doit le montrer plutôt que de le cacher."""
    formules = _formules(classeur)
    assert any("$J$1*C" in f and "/250" in f for f in formules)


def test_aucune_macro_n_est_pretendue(classeur):
    """Un script ne peut pas écrire de macro. Le fichier porte donc l'extension des classeurs sans
    macro, et rien dans l'archive ne prétend le contraire."""
    assert classeur.suffix == ".xlsx"
    assert not any("vbaProject" in n for n in zipfile.ZipFile(classeur).namelist())
