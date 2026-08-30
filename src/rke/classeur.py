"""Le classeur Excel des tests de dépassement, à formules vivantes.

**Pourquoi un classeur et pas seulement un tableau.** Dans un service de risque, personne ne relit
du Python. On relit un classeur, on change une case, et on regarde ce que le résultat devient. Un
tableau de chiffres figés ne permet pas cela ; un classeur à formules oui.

**Ce qu'il contient.** Une feuille par test, et dans chaque feuille les formules sont écrites en
Excel, pas calculées en Python puis collées. Le lecteur peut donc changer le nombre de jours, le
nombre de dépassements ou le seuil, et voir la statistique et sa valeur p bouger sous ses yeux.

**Ce qu'il ne contient pas, et pourquoi.** Aucune macro. Un script ne peut pas en écrire : mesuré le
2026-08-30, un fichier enregistré sous l'extension des classeurs à macros ne contient aucun projet
de macros, et Excel l'ouvre comme un classeur ordinaire. Le module `vba/KupiecTest.bas` livre donc
le même calcul en Visual Basic, sous forme de texte à importer en deux clics, et le classeur reste
utilisable sans lui.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITRE = Font(bold=True, color="FFFFFF")
FOND = PatternFill("solid", fgColor="0072B2")
GRAS = Font(bold=True)
CADRE = Border(bottom=Side(style="thin", color="B0B0B0"))
POURCENT = "0.00%"
DECIMAL = "0.0000"


def _entete(feuille, titres: list[str], ligne: int = 1) -> None:
    for colonne, texte in enumerate(titres, start=1):
        case = feuille.cell(row=ligne, column=colonne, value=texte)
        case.font, case.fill = TITRE, FOND
        case.alignment = Alignment(horizontal="center", wrap_text=True)
    feuille.row_dimensions[ligne].height = 30


def _largeurs(feuille, largeurs: list[int]) -> None:
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur


def _notice(feuille, ligne: int, lignes: list[str]) -> int:
    for i, texte in enumerate(lignes):
        case = feuille.cell(row=ligne + i, column=1, value=texte)
        case.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 0:
            case.font = GRAS
    return ligne + len(lignes) + 1


def feuille_kupiec(classeur: Workbook, backtests: pd.DataFrame, niveau: float = 0.99) -> None:
    """Le test de Kupiec, écrit en formules Excel plutôt qu'en résultats collés.

    Le test répond à une question simple : le modèle annonce qu'il sera dépassé un jour sur cent ;
    l'a-t-il été à peu près une fois sur cent ? La statistique compare la fréquence observée à la
    fréquence promise, et sa valeur p dit si l'écart tient dans le hasard.
    """
    feuille = classeur.create_sheet("Kupiec")
    _largeurs(feuille, [26, 12, 14, 12, 13, 15, 15, 13, 12])
    _entete(feuille, ["Modèle", "Jours", "Dépassements", "Attendus", "Taux observé",
                      "Vraisemblance sous le modèle", "Vraisemblance libre",
                      "Statistique", "Valeur p"])

    feuille["K1"], feuille["L1"] = "Niveau de confiance", niveau
    feuille["L1"].number_format = POURCENT
    feuille["K1"].font = GRAS

    for i, (nom, ligne) in enumerate(backtests.iterrows(), start=2):
        p = "$L$1"
        feuille.cell(row=i, column=1, value=str(nom))
        feuille.cell(row=i, column=2, value=int(ligne["n_jours"]))
        feuille.cell(row=i, column=3, value=int(ligne["depassements"]))
        feuille.cell(row=i, column=4, value=f"=B{i}*(1-{p})").number_format = "0.0"
        feuille.cell(row=i, column=5, value=f"=C{i}/B{i}").number_format = "0.000%"
        # la vraisemblance sous le taux promis, puis sous le taux réellement observé
        feuille.cell(row=i, column=6,
                     value=f"=(B{i}-C{i})*LN({p})+C{i}*LN(1-{p})").number_format = DECIMAL
        feuille.cell(row=i, column=7,
                     value=f"=IF(C{i}=0,B{i}*LN(1-E{i}),"
                           f"(B{i}-C{i})*LN(1-E{i})+C{i}*LN(E{i}))").number_format = DECIMAL
        feuille.cell(row=i, column=8, value=f"=-2*(F{i}-G{i})").number_format = DECIMAL
        # CHIDIST et non CHISQ.DIST.RT : les fonctions ajoutées après Excel 2007 doivent être
        # écrites avec un préfixe technique dans le fichier, faute de quoi Excel affiche une erreur
        # de nom. La forme ancienne évite le piège et fonctionne dans toutes les versions.
        feuille.cell(row=i, column=9, value=f"=CHIDIST(H{i},1)").number_format = "0.0000"
        feuille.cell(row=i, column=9).font = GRAS
        for colonne in range(1, 10):
            feuille.cell(row=i, column=colonne).border = CADRE

    _notice(feuille, len(backtests) + 3, [
        "Comment lire cette feuille",
        "Chaque ligne est un modèle de valeur à risque. La colonne « Dépassements » compte les jours "
        "où la perte a dépassé ce que le modèle annonçait.",
        "La statistique compare deux vraisemblances : celle du taux promis et celle du taux "
        "réellement observé. Plus l'écart est grand, plus la statistique monte.",
        "La valeur p est la probabilité d'observer un écart au moins aussi grand si le modèle disait "
        "vrai. Sous 5 %, le modèle est rejeté.",
        "Changez le niveau de confiance en L1, ou le nombre de dépassements en colonne C : tout se "
        "recalcule, car rien n'est collé.",
    ])


def feuille_feux(classeur: Workbook, feux: dict[str, pd.DataFrame]) -> None:
    """Les zones de Bâle année par année, avec les seuils écrits en formule.

    Le régulateur classe une année en vert, jaune ou rouge selon le nombre de dépassements. Les
    seuils de 5 et 10 valent pour 250 jours d'observation ; ils sont mis à l'échelle du nombre de
    jours réellement observés, et cette mise à l'échelle est visible dans la formule.
    """
    feuille = classeur.create_sheet("Feux de Bâle")
    _largeurs(feuille, [10, 22, 10, 15, 14, 14, 12])
    _entete(feuille, ["Année", "Modèle", "Jours", "Dépassements", "Seuil jaune", "Seuil rouge",
                      "Zone"])

    feuille["I1"], feuille["J1"] = "Seuil jaune sur 250 jours", 5
    feuille["I2"], feuille["J2"] = "Seuil rouge sur 250 jours", 10
    feuille["I1"].font = feuille["I2"].font = GRAS

    ligne = 2
    for nom, table in feux.items():
        # l'année est l'index de ces tables, pas une colonne : la lire comme une colonne échoue
        for annee, donnees in table.iterrows():
            feuille.cell(row=ligne, column=1, value=int(annee))
            feuille.cell(row=ligne, column=2, value=nom)
            feuille.cell(row=ligne, column=3, value=int(donnees["n_jours"]))
            feuille.cell(row=ligne, column=4, value=int(donnees["depassements"]))
            feuille.cell(row=ligne, column=5,
                         value=f"=$J$1*C{ligne}/250").number_format = "0.0"
            feuille.cell(row=ligne, column=6,
                         value=f"=$J$2*C{ligne}/250").number_format = "0.0"
            feuille.cell(row=ligne, column=7,
                         value=f'=IF(D{ligne}>=F{ligne},"rouge",'
                               f'IF(D{ligne}>=E{ligne},"jaune","vert"))')
            ligne += 1

    _notice(feuille, ligne + 1, [
        "Comment lire cette feuille",
        "Le régulateur compte les dépassements de chaque année et range l'année en vert, jaune ou "
        "rouge. Une année rouge coûte du capital en plus.",
        "Les seuils de 5 et 10 valent pour 250 jours de bourse. Ils sont ici mis à l'échelle du "
        "nombre de jours réellement observés dans l'année, ce que la formule montre.",
        "Changez les seuils en J1 et J2 pour voir combien d'années changent de couleur.",
    ])


def feuille_notice(classeur: Workbook) -> None:
    """La première feuille : ce que le classeur contient et ce qu'il ne contient pas."""
    feuille = classeur.create_sheet("Notice", 0)
    _largeurs(feuille, [110])
    ligne = _notice(feuille, 1, [
        "Classeur des tests de dépassement",
        "Ce classeur accompagne le dépôt 06-risque-marche. Il contient deux feuilles, et dans les "
        "deux les formules sont écrites en Excel : rien n'est collé depuis Python.",
        "",
    ])
    ligne = _notice(feuille, ligne, [
        "Feuille « Kupiec »",
        "Le test qui demande si un modèle annonçant un dépassement sur cent jours a bien été dépassé "
        "environ une fois sur cent.",
        "",
    ])
    ligne = _notice(feuille, ligne, [
        "Feuille « Feux de Bâle »",
        "Le classement réglementaire de chaque année en vert, jaune ou rouge, selon le nombre de "
        "dépassements.",
        "",
    ])
    _notice(feuille, ligne, [
        "Pourquoi il n'y a pas de macro",
        "Une macro ne se fabrique pas par script : un fichier enregistré sous l'extension des "
        "classeurs à macros ne contient aucun projet de macros, et Excel l'ouvre comme un classeur "
        "ordinaire. C'est mesuré, pas supposé.",
        "Le même test est donc livré en Visual Basic dans le fichier vba/KupiecTest.bas du dépôt. "
        "Pour l'utiliser : ouvrir l'éditeur Visual Basic par Alt+F11, menu Fichier, Importer un "
        "fichier, choisir KupiecTest.bas. La fonction KUPIEC(jours; depassements; niveau) devient "
        "alors disponible dans les cellules.",
        "Le classeur fonctionne entièrement sans cette macro : les formules Excel font le même "
        "calcul.",
    ])


def construire(backtests: pd.DataFrame, feux: dict[str, pd.DataFrame], dest: Path) -> Path:
    """Le classeur complet, écrit sur le disque."""
    classeur = Workbook()
    classeur.remove(classeur.active)
    feuille_kupiec(classeur, backtests)
    feuille_feux(classeur, feux)
    feuille_notice(classeur)
    dest.parent.mkdir(parents=True, exist_ok=True)
    classeur.save(dest)
    return dest
